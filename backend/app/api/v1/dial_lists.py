"""
Dial Lists API Endpoints

CRUD for ViciDial-style dial lists, entries, DNC, campaign-list links,
hopper, and dispositions. Includes Excel/CSV file upload with formula
sanitization.
"""

import csv
import io
import logging
import re
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from sqlalchemy.orm import Session
from sqlalchemy import or_, text

from app.core.database import get_db
from app.core.config import settings
from app.models.models import (
    DialList, DialListEntry, DialAttempt, DNCList,
    CampaignList, DialHopper, CampaignDisposition,
    Campaign, DialListStatus, DialEntryStatus,
    Agent,
)
from app.models import User
from app.api.v1.auth import get_current_user
from app.schemas.schemas import (
    DialListCreate, DialListUpdate, DialListResponse,
    DialListEntryCreate, DialListEntryUpdate,
    DialListEntryBulkCreate, DialListEntryResponse,
    DialAttemptResponse,
    DNCListCreate, DNCListResponse,
    CampaignListCreate, CampaignListResponse,
    DialHopperResponse,
    CampaignDispositionCreate, CampaignDispositionResponse,
    ExcelUploadResponse,
    DuplicateCheckResponse, DuplicateDetail, DuplicateListInfo,
    PaginatedResponse,
)

logger = logging.getLogger(__name__)

# Maximum upload size in bytes (from settings, default 10 MB)
MAX_UPLOAD_BYTES = getattr(settings, "MAX_UPLOAD_SIZE_MB", 10) * 1024 * 1024

# Characters that indicate Excel formula injection
# Note: '+' is NOT included because it's legitimate in phone numbers
# (E.164 format) and poses minimal injection risk.
_FORMULA_PREFIXES = ("=", "-", "@", "\t", "\r")


def _sanitize_cell(value: str) -> str:
    """Strip leading formula-injection characters from a cell value."""
    if not value:
        return value
    stripped = value.strip()
    while stripped and stripped[0] in _FORMULA_PREFIXES:
        stripped = stripped[1:].lstrip()
    return stripped


def _cell_to_str(cell) -> str:
    """Convert an openpyxl cell value to a clean string.

    Excel stores numbers as IEEE 754 doubles.  Phone numbers like
    905551234567 come back as ``float`` 905551234567.0.  A naive
    ``str()`` would yield ``'905551234567.0'`` which corrupts the
    phone after digit-only stripping.  This helper converts
    whole-number floats to ``int`` first so the string is clean.

    Also handles scientific notation strings like ``'4.92167E+12'``
    which may appear when data is exported through certain tools.
    """
    if cell is None:
        return ""
    if isinstance(cell, (int, float)):
        try:
            f = float(cell)
            if f == f and f == int(f):  # NaN check + whole check
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    # Handle scientific notation strings (e.g. "4.92167E+12")
    s = str(cell)
    if "e+" in s.lower() or "e-" in s.lower():
        try:
            f = float(s)
            if f == f and f == int(f):
                return str(int(f))
        except (ValueError, OverflowError):
            pass
    return s


def _normalize_phone(raw: str, default_country_code: str = "") -> Optional[str]:
    """
    Normalize a phone number to digits-only.
    Returns None if the number is clearly invalid (too short/long).
    Accepts E.164 (+905551234567) or local (5551234567) formats.
    """
    if not raw:
        logger.warning("_normalize_phone: empty input")
        return None
    cleaned = raw.strip()
    # Remember if there was a leading +
    has_plus = cleaned.startswith("+")
    # Remove everything except digits
    pure_digits = re.sub(r"\D", "", cleaned)
    if len(pure_digits) < 7 or len(pure_digits) > 15:
        logger.warning(
            "_normalize_phone REJECTED: raw=%r -> pure_digits=%r "
            "(len=%d, need 7-15)",
            raw, pure_digits, len(pure_digits),
        )
        return None
    # Reconstruct with optional leading +
    digits = ("+" + pure_digits) if has_plus else pure_digits
    # Add country code if not present and a default is provided
    if default_country_code and not has_plus:
        digits = default_country_code + pure_digits
    return digits


def _parse_excel_rows(contents: bytes, filename: str, column_mapping: dict) -> List[dict]:
    """
    Parse Excel (.xlsx/.xls) or CSV file into a list of row dicts.
    column_mapping maps field names to column letters/indices.
    """
    rows: List[dict] = []

    if filename.endswith(".csv"):
        text_content = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text_content))
        header = next(reader, None)
        if not header:
            return rows

        # Build column index map from header names or column letters
        col_indices = _resolve_column_indices(header, column_mapping)

        for csv_row in reader:
            row_data = _extract_row(csv_row, col_indices)
            if row_data:
                rows.append(row_data)
    else:
        # Excel file - use openpyxl
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(
                status_code=500,
                detail="openpyxl is required for Excel file processing",
            )

        wb = openpyxl.load_workbook(io.BytesIO(contents), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            return rows

        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            logger.warning("Excel file has no rows at all")
            return rows

        header = [str(c) if c else "" for c in all_rows[0]]
        col_indices = _resolve_column_indices(header, column_mapping)

        logger.warning(
            "UPLOAD DEBUG: header=%r, column_mapping=%r, "
            "resolved_indices=%r, total_data_rows=%d",
            header, column_mapping, col_indices, len(all_rows) - 1,
        )

        for row_num, excel_row in enumerate(all_rows[1:], start=2):
            cell_values = [_cell_to_str(c) for c in excel_row]

            # Log every row's phone cell for debugging
            phone_idx = col_indices.get("phone_number")
            raw_phone_cell = None
            if phone_idx is not None and phone_idx < len(excel_row):
                raw_phone_cell = excel_row[phone_idx]
            logger.warning(
                "UPLOAD DEBUG row %d: raw_phone_cell=%r (type=%s), "
                "cell_to_str=%r, all_cells=%r",
                row_num,
                raw_phone_cell,
                type(raw_phone_cell).__name__,
                cell_values[phone_idx] if phone_idx is not None and phone_idx < len(cell_values) else "NO_IDX",
                cell_values,
            )

            row_data = _extract_row(cell_values, col_indices)
            if row_data:
                rows.append(row_data)
            else:
                logger.warning(
                    "UPLOAD DEBUG row %d SKIPPED: raw_excel_cells=%r",
                    row_num,
                    [c for c in excel_row],
                )

        wb.close()

    return rows


def _resolve_column_indices(header: List[str], mapping: dict) -> dict:
    """
    Resolve column mapping to numeric indices.
    Mapping values can be column letters (A, B, ...) or header names.
    """
    field_map = {
        "phone_column": "phone_number",
        "first_name_column": "first_name",
        "last_name_column": "last_name",
        "title_column": "title",
        "email_column": "email",
        "company_column": "company",
        "address_column": "address",
        "timezone_column": "timezone",
        "notes_column": "notes",
    }

    indices: dict = {}
    header_lower = [h.lower().strip() for h in header]

    for map_key, field_name in field_map.items():
        col_ref = mapping.get(map_key)
        if not col_ref:
            continue

        col_ref_str = str(col_ref).strip()

        # Try as column letter (A=0, B=1, ...)
        if len(col_ref_str) == 1 and col_ref_str.upper().isalpha():
            idx = ord(col_ref_str.upper()) - ord("A")
            if 0 <= idx < len(header):
                indices[field_name] = idx
                continue

        # Try as column number (1-based)
        if col_ref_str.isdigit():
            idx = int(col_ref_str) - 1
            if 0 <= idx < len(header):
                indices[field_name] = idx
                continue

        # Try as header name match
        col_lower = col_ref_str.lower()
        for i, h in enumerate(header_lower):
            if h == col_lower:
                indices[field_name] = i
                break

    return indices


def _extract_row(cells: List[str], col_indices: dict) -> Optional[dict]:
    """Extract a row dict from cell values using resolved column indices."""
    phone_idx = col_indices.get("phone_number")
    if phone_idx is None or phone_idx >= len(cells):
        return None

    phone_raw = _sanitize_cell(cells[phone_idx])
    phone = _normalize_phone(phone_raw)
    if not phone:
        logger.warning("Row skipped: phone_raw=%r -> normalized=None", phone_raw)
        return None

    row: dict = {"phone_number": phone}
    for field, idx in col_indices.items():
        if field == "phone_number" or idx >= len(cells):
            continue
        row[field] = _sanitize_cell(cells[idx])

    return row


# ============================================================
# Dial Lists CRUD
# ============================================================

router = APIRouter(prefix="/dial-lists", tags=["Dial Lists"])


@router.get("/", response_model=PaginatedResponse)
async def list_dial_lists(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> PaginatedResponse:
    """List all dial lists with pagination and filtering."""
    query = db.query(DialList).filter(DialList.status != DialListStatus.ARCHIVED)

    if status:
        query = query.filter(DialList.status == status)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(DialList.name.ilike(pattern), DialList.description.ilike(pattern))
        )

    total = query.count()
    items = (
        query.order_by(DialList.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PaginatedResponse(
        items=[DialListResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.post("/", response_model=DialListResponse, status_code=201)
async def create_dial_list(
    data: DialListCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DialListResponse:
    """Create a new dial list."""
    # Validate agent_id if provided
    if data.agent_id:
        agent = db.get(Agent, data.agent_id)
        if not agent:
            raise HTTPException(status_code=404, detail="Agent not found")

    dial_list = DialList(
        name=data.name,
        description=data.description,
        agent_id=data.agent_id,
        owner_id=current_user.id,
    )
    db.add(dial_list)
    db.commit()
    db.refresh(dial_list)
    return DialListResponse.model_validate(dial_list)


@router.post("/preview-headers")
async def preview_file_headers(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
) -> dict:
    """
    Preview Excel/CSV file headers and sample rows.
    Returns column headers and first 3 data rows for the column mapping UI.
    """
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx, .xls, and .csv files are supported",
        )

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    headers: List[str] = []
    sample_rows: List[List[str]] = []

    if filename.endswith(".csv"):
        text_content = contents.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text_content))
        header_row = next(reader, None)
        if header_row:
            headers = [h.strip() for h in header_row if h and h.strip()]
            for i, row in enumerate(reader):
                if i >= 3:
                    break
                sample_rows.append(
                    [str(c).strip() if c else "" for c in row[: len(headers)]]
                )
    else:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(status_code=500, detail="openpyxl required")

        wb = openpyxl.load_workbook(
            io.BytesIO(contents), read_only=True, data_only=True
        )
        ws = wb.active
        if ws:
            all_rows = list(ws.iter_rows(values_only=True))
            if all_rows:
                headers = [
                    str(c).strip() if c else f"Column {i + 1}"
                    for i, c in enumerate(all_rows[0])
                ]
                for row in all_rows[1:4]:
                    sample_rows.append(
                        [
                            _cell_to_str(c).strip() if c is not None else ""
                            for c in row[: len(headers)]
                        ]
                    )
        wb.close()

    if not headers:
        raise HTTPException(status_code=400, detail="No headers found in file")

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "total_columns": len(headers),
    }


@router.get("/{list_id}", response_model=DialListResponse)
async def get_dial_list(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DialListResponse:
    """Get a specific dial list with live entry counts."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    # Refresh counters from actual entries
    base_q = db.query(DialListEntry).filter(DialListEntry.list_id == list_id)
    dial_list.total_numbers = base_q.count()
    dial_list.active_numbers = base_q.filter(
        DialListEntry.status.in_([DialEntryStatus.NEW, DialEntryStatus.CALLBACK])
    ).count()
    dial_list.completed_numbers = base_q.filter(
        DialListEntry.status == DialEntryStatus.COMPLETED
    ).count()
    dial_list.invalid_numbers = base_q.filter(
        DialListEntry.status == DialEntryStatus.INVALID
    ).count()
    db.commit()

    return DialListResponse.model_validate(dial_list)


@router.put("/{list_id}", response_model=DialListResponse)
async def update_dial_list(
    list_id: int,
    data: DialListUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DialListResponse:
    """Update a dial list name/description/status."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(dial_list, key, value)

    db.commit()
    db.refresh(dial_list)
    return DialListResponse.model_validate(dial_list)


@router.delete("/{list_id}")
async def delete_dial_list(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Soft-delete a dial list by setting status to archived."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    dial_list.status = DialListStatus.ARCHIVED
    db.commit()
    return {"message": "Dial list archived successfully", "id": list_id}


# ============================================================
# Check Duplicates (Preview)
# ============================================================

@router.post("/{list_id}/check-duplicates", response_model=DuplicateCheckResponse)
async def check_duplicates(
    list_id: int,
    file: UploadFile = File(...),
    phone_column: str = Form("A"),
    first_name_column: Optional[str] = Form(None),
    last_name_column: Optional[str] = Form(None),
    title_column: Optional[str] = Form(None),
    email_column: Optional[str] = Form(None),
    company_column: Optional[str] = Form(None),
    address_column: Optional[str] = Form(None),
    timezone_column: Optional[str] = Form(None),
    notes_column: Optional[str] = Form(None),
    country_code: str = Form(""),
    duplicate_mode: str = Form("file_only"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DuplicateCheckResponse:
    """
    Check for duplicate phone numbers BEFORE actually uploading.

    duplicate_mode:
        - file_only: only detect duplicates within the uploaded file itself
        - same_agent: also check numbers in other lists assigned to the same agent
        - all_system: check numbers across ALL lists in the entire system
    """
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, and .csv files are supported")

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    # Build column mapping
    column_mapping = {"phone_column": phone_column}
    if first_name_column:
        column_mapping["first_name_column"] = first_name_column
    if last_name_column:
        column_mapping["last_name_column"] = last_name_column
    if title_column:
        column_mapping["title_column"] = title_column
    if email_column:
        column_mapping["email_column"] = email_column
    if company_column:
        column_mapping["company_column"] = company_column
    if address_column:
        column_mapping["address_column"] = address_column
    if timezone_column:
        column_mapping["timezone_column"] = timezone_column
    if notes_column:
        column_mapping["notes_column"] = notes_column

    try:
        rows = _parse_excel_rows(contents, filename, column_mapping)
    except Exception as e:
        logger.error(f"Error parsing file for duplicate check: {e}")
        raise HTTPException(status_code=400, detail="Failed to parse file")

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found in file")

    # Normalize all phone numbers
    phones: list[str] = []
    for row_data in rows:
        phone = row_data["phone_number"]
        if country_code:
            normalized = _normalize_phone(phone, country_code)
            if normalized:
                phone = normalized
        phones.append(phone)

    total_rows = len(phones)

    # --- Step 1: Within-file duplicates ---
    seen: set[str] = set()
    file_dup_phones: set[str] = set()
    for p in phones:
        if p in seen:
            file_dup_phones.add(p)
        else:
            seen.add(p)

    unique_phones = seen  # All unique numbers in file

    # --- Step 2: System duplicates based on mode ---
    system_dup_details: list[DuplicateDetail] = []
    system_numbers_map: dict[str, list[DuplicateListInfo]] = {}

    if duplicate_mode in ("same_agent", "all_system"):
        # Build query for existing numbers
        entry_q = db.query(
            DialListEntry.phone_number,
            DialList.id,
            DialList.name,
        ).join(DialList, DialListEntry.list_id == DialList.id).filter(
            DialList.status != DialListStatus.ARCHIVED,
            DialListEntry.phone_number.in_(list(unique_phones)),
        )

        if duplicate_mode == "same_agent" and dial_list.agent_id:
            # Only check lists with the same agent
            entry_q = entry_q.filter(DialList.agent_id == dial_list.agent_id)

        for phone, lid, lname in entry_q.all():
            if phone not in system_numbers_map:
                system_numbers_map[phone] = []
            system_numbers_map[phone].append(
                DuplicateListInfo(list_id=lid, list_name=lname)
            )

    # Build duplicate details
    duplicates: list[DuplicateDetail] = []

    # File duplicates
    for p in sorted(file_dup_phones):
        duplicates.append(DuplicateDetail(
            phone=p,
            source="file",
            found_in_lists=[],
        ))

    # System duplicates
    for p, lists_info in sorted(system_numbers_map.items()):
        duplicates.append(DuplicateDetail(
            phone=p,
            source="system",
            found_in_lists=lists_info,
        ))

    all_dup_phones = file_dup_phones | set(system_numbers_map.keys())
    clean_count = len(unique_phones) - len(unique_phones & set(system_numbers_map.keys()))

    return DuplicateCheckResponse(
        total_rows=total_rows,
        unique_numbers=len(unique_phones),
        file_duplicates=len(file_dup_phones),
        system_duplicates=len(system_numbers_map),
        duplicates=duplicates,
        clean_count=clean_count,
    )


# ============================================================
# Excel / CSV Upload
# ============================================================

@router.post("/{list_id}/upload", response_model=ExcelUploadResponse)
async def upload_file(
    list_id: int,
    file: UploadFile = File(...),
    phone_column: str = Form("A"),
    first_name_column: Optional[str] = Form(None),
    last_name_column: Optional[str] = Form(None),
    title_column: Optional[str] = Form(None),
    email_column: Optional[str] = Form(None),
    company_column: Optional[str] = Form(None),
    address_column: Optional[str] = Form(None),
    timezone_column: Optional[str] = Form(None),
    notes_column: Optional[str] = Form(None),
    country_code: str = Form(""),
    duplicate_mode: str = Form("file_only"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ExcelUploadResponse:
    """
    Upload an Excel (.xlsx, .xls) or CSV file to populate a dial list.

    Column mapping tells the system which spreadsheet columns map to
    which fields. Values can be column letters (A, B, ...) or header names.

    duplicate_mode controls how duplicates are detected:
        - none: no duplicate checking at all
        - file_only: only within-file and same-list duplicates
        - same_agent: also skip numbers found in other lists of the same agent
        - all_system: skip numbers found in any list across the system
    """
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    # Validate file type
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx, .xls, and .csv files are supported",
        )

    # Read file contents with size check
    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File too large. Maximum size is {settings.MAX_UPLOAD_SIZE_MB} MB",
        )

    # Build column mapping from form fields
    column_mapping = {"phone_column": phone_column}
    if first_name_column:
        column_mapping["first_name_column"] = first_name_column
    if last_name_column:
        column_mapping["last_name_column"] = last_name_column
    if title_column:
        column_mapping["title_column"] = title_column
    if email_column:
        column_mapping["email_column"] = email_column
    if company_column:
        column_mapping["company_column"] = company_column
    if address_column:
        column_mapping["address_column"] = address_column
    if timezone_column:
        column_mapping["timezone_column"] = timezone_column
    if notes_column:
        column_mapping["notes_column"] = notes_column

    # Parse file
    try:
        rows = _parse_excel_rows(contents, filename, column_mapping)
    except Exception as e:
        logger.error(f"Error parsing uploaded file: {e}")
        raise HTTPException(status_code=400, detail="Failed to parse file. Check the file format.")

    if not rows:
        raise HTTPException(status_code=400, detail="No valid rows found in file")

    logger.warning(
        "UPLOAD LOOP DEBUG: parsed_rows=%d, duplicate_mode=%s, list_id=%d",
        len(rows), duplicate_mode, list_id,
    )

    # Pre-load DNC numbers for fast lookup
    dnc_numbers = set(row[0] for row in db.query(DNCList.phone_number).all())

    # Pre-load existing numbers in this list (skipped for 'none' mode)
    existing_numbers: set = set()
    if duplicate_mode != "none":
        existing_numbers = set(
            row[0]
            for row in db.query(DialListEntry.phone_number)
            .filter(DialListEntry.list_id == list_id)
            .all()
        )

    # Extended duplicate check based on mode
    if duplicate_mode in ("same_agent", "all_system"):
        ext_q = db.query(DialListEntry.phone_number).join(
            DialList, DialListEntry.list_id == DialList.id
        ).filter(
            DialList.status != DialListStatus.ARCHIVED,
            DialList.id != list_id,
        )
        if duplicate_mode == "same_agent" and dial_list.agent_id:
            ext_q = ext_q.filter(DialList.agent_id == dial_list.agent_id)
        for (phone,) in ext_q.all():
            existing_numbers.add(phone)

    logger.warning(
        "UPLOAD LOOP DEBUG: dnc_count=%d, pre_existing_count=%d",
        len(dnc_numbers), len(existing_numbers),
    )

    success = 0
    errors = 0
    duplicates = 0
    dnc_count = 0
    error_details: List[dict] = []
    batch: List[dict] = []
    batch_size = 1000

    for idx, row_data in enumerate(rows):
        phone = row_data["phone_number"]

        # Apply country code normalization
        if country_code:
            normalized = _normalize_phone(phone, country_code)
            if normalized:
                phone = normalized
                row_data["phone_number"] = phone

        # DNC check
        if phone in dnc_numbers:
            dnc_count += 1
            errors += 1
            error_details.append({"row": idx + 2, "phone": phone, "reason": "DNC listed"})
            logger.warning("UPLOAD LOOP row %d DNC: phone=%r", idx + 2, phone)
            continue

        # Duplicate check
        if phone in existing_numbers:
            duplicates += 1
            logger.warning("UPLOAD LOOP row %d DUP: phone=%r", idx + 2, phone)
            continue

        # Build custom_fields for extra columns (title, address, etc.)
        extra_fields: dict = {}
        if row_data.get("title"):
            extra_fields["title"] = row_data["title"]
        if row_data.get("address"):
            extra_fields["address"] = row_data["address"]

        # Build mapping for bulk insert
        batch.append({
            "list_id": list_id,
            "phone_number": phone,
            "first_name": row_data.get("first_name", ""),
            "last_name": row_data.get("last_name", ""),
            "email": row_data.get("email", ""),
            "company": row_data.get("company", ""),
            "timezone": row_data.get("timezone", ""),
            "notes": row_data.get("notes", ""),
            "priority": 0,
            "status": DialEntryStatus.NEW,
            "call_attempts": 0,
            "max_attempts": 3,
            "dnc_flag": False,
            "custom_fields": extra_fields,
        })
        existing_numbers.add(phone)
        success += 1

        # Flush batch every batch_size rows
        if len(batch) >= batch_size:
            db.bulk_insert_mappings(DialListEntry, batch)
            batch.clear()

    # Flush remaining
    if batch:
        db.bulk_insert_mappings(DialListEntry, batch)

    # Update list counters
    dial_list.total_numbers += success
    dial_list.active_numbers += success

    db.commit()

    logger.warning(
        "UPLOAD RESULT: total=%d, success=%d, errors=%d, "
        "duplicates=%d, dnc=%d",
        len(rows), success, errors, duplicates, dnc_count,
    )

    return ExcelUploadResponse(
        total=len(rows),
        success=success,
        errors=errors,
        duplicates=duplicates,
        error_details=error_details[:100] if error_details else None,
    )


# ============================================================
# Dial List Entries
# ============================================================

@router.get("/{list_id}/entries", response_model=PaginatedResponse)
async def list_entries(
    list_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> PaginatedResponse:
    """List entries in a dial list with filtering."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    query = db.query(DialListEntry).filter(DialListEntry.list_id == list_id)

    if status:
        query = query.filter(DialListEntry.status == status)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            or_(
                DialListEntry.phone_number.ilike(pattern),
                DialListEntry.first_name.ilike(pattern),
                DialListEntry.last_name.ilike(pattern),
                DialListEntry.company.ilike(pattern),
            )
        )

    total = query.count()
    items = (
        query.order_by(DialListEntry.priority.desc(), DialListEntry.id)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PaginatedResponse(
        items=[DialListEntryResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@router.post("/{list_id}/entries", response_model=DialListEntryResponse, status_code=201)
async def add_entry(
    list_id: int,
    data: DialListEntryCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DialListEntryResponse:
    """Add a single entry to a dial list."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    # DNC check
    dnc = db.query(DNCList).filter(DNCList.phone_number == data.phone_number).first()
    if dnc:
        raise HTTPException(status_code=409, detail="Phone number is on the DNC list")

    entry = DialListEntry(
        list_id=list_id,
        phone_number=data.phone_number,
        first_name=data.first_name,
        last_name=data.last_name,
        email=data.email,
        company=data.company,
        timezone=data.timezone,
        priority=data.priority,
        max_attempts=data.max_attempts,
        custom_fields=data.custom_fields or {},
        notes=data.notes,
    )
    db.add(entry)
    dial_list.total_numbers += 1
    dial_list.active_numbers += 1
    db.commit()
    db.refresh(entry)
    return DialListEntryResponse.model_validate(entry)


@router.post("/{list_id}/entries/bulk", response_model=ExcelUploadResponse, status_code=201)
async def bulk_add_entries(
    list_id: int,
    data: DialListEntryBulkCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ExcelUploadResponse:
    """Bulk add entries to a dial list via JSON body."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    dnc_numbers = set(row[0] for row in db.query(DNCList.phone_number).all())
    existing_numbers = set(
        row[0]
        for row in db.query(DialListEntry.phone_number)
        .filter(DialListEntry.list_id == list_id)
        .all()
    )

    success = 0
    errors = 0
    duplicates = 0
    error_details: List[dict] = []

    for idx, entry_data in enumerate(data.entries):
        phone = entry_data.phone_number
        if phone in dnc_numbers:
            errors += 1
            error_details.append({"row": idx + 1, "phone": phone, "reason": "DNC listed"})
            continue
        if phone in existing_numbers:
            duplicates += 1
            continue

        entry = DialListEntry(
            list_id=list_id,
            phone_number=phone,
            first_name=entry_data.first_name,
            last_name=entry_data.last_name,
            email=entry_data.email,
            company=entry_data.company,
            timezone=entry_data.timezone,
            priority=entry_data.priority,
            max_attempts=entry_data.max_attempts,
            custom_fields=entry_data.custom_fields or {},
            notes=entry_data.notes,
        )
        db.add(entry)
        existing_numbers.add(phone)
        success += 1

    dial_list.total_numbers += success
    dial_list.active_numbers += success
    db.commit()

    return ExcelUploadResponse(
        total=len(data.entries),
        success=success,
        errors=errors,
        duplicates=duplicates,
        error_details=error_details if error_details else None,
    )


@router.patch("/{list_id}/entries/{entry_id}", response_model=DialListEntryResponse)
async def update_entry(
    list_id: int,
    entry_id: int,
    data: DialListEntryUpdate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DialListEntryResponse:
    """Update a dial list entry (status, notes, priority, callback time)."""
    entry = (
        db.query(DialListEntry)
        .filter(DialListEntry.id == entry_id, DialListEntry.list_id == list_id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    for key, value in data.model_dump(exclude_unset=True).items():
        setattr(entry, key, value)

    db.commit()
    db.refresh(entry)
    return DialListEntryResponse.model_validate(entry)


@router.delete("/{list_id}/entries/{entry_id}")
async def delete_entry(
    list_id: int,
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Delete a single entry from a dial list."""
    entry = (
        db.query(DialListEntry)
        .filter(DialListEntry.id == entry_id, DialListEntry.list_id == list_id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")

    dial_list = db.get(DialList, list_id)
    if dial_list:
        dial_list.total_numbers = max(0, dial_list.total_numbers - 1)
        if entry.status in (DialEntryStatus.NEW, DialEntryStatus.CALLBACK):
            dial_list.active_numbers = max(0, dial_list.active_numbers - 1)

    db.delete(entry)
    db.commit()
    return {"message": "Entry deleted successfully", "id": entry_id}


@router.post("/{list_id}/reset")
async def reset_entries(
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Reset all entries in a dial list to NEW status for re-dialing."""
    dial_list = db.get(DialList, list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    # Reset non-DNC entries back to NEW
    updated = (
        db.query(DialListEntry)
        .filter(
            DialListEntry.list_id == list_id,
            DialListEntry.dnc_flag == False,
            DialListEntry.status != DialEntryStatus.DNC,
        )
        .update({
            "status": DialEntryStatus.NEW,
            "call_attempts": 0,
            "last_attempt_at": None,
            "next_callback_at": None,
        })
    )

    # Refresh counters
    total = db.query(DialListEntry).filter(DialListEntry.list_id == list_id).count()
    active = (
        db.query(DialListEntry)
        .filter(
            DialListEntry.list_id == list_id,
            DialListEntry.status.in_([DialEntryStatus.NEW, DialEntryStatus.CALLBACK]),
        )
        .count()
    )
    dial_list.total_numbers = total
    dial_list.active_numbers = active
    dial_list.completed_numbers = 0

    db.commit()
    return {"message": f"Reset {updated} entries to NEW status", "reset_count": updated}


# ============================================================
# Dial Attempts
# ============================================================

@router.get("/entries/{entry_id}/attempts", response_model=List[DialAttemptResponse])
async def list_attempts(
    entry_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[DialAttemptResponse]:
    """List all dial attempts for an entry."""
    attempts = (
        db.query(DialAttempt)
        .filter(DialAttempt.entry_id == entry_id)
        .order_by(DialAttempt.attempt_number)
        .all()
    )
    return [DialAttemptResponse.model_validate(a) for a in attempts]


# ============================================================
# DNC List
# ============================================================

dnc_router = APIRouter(prefix="/dnc", tags=["DNC List"])


@dnc_router.get("/", response_model=PaginatedResponse)
async def list_dnc(
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    search: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> PaginatedResponse:
    """List all DNC entries."""
    query = db.query(DNCList)
    if search:
        query = query.filter(DNCList.phone_number.ilike(f"%{search}%"))

    total = query.count()
    items = (
        query.order_by(DNCList.created_at.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PaginatedResponse(
        items=[DNCListResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )


@dnc_router.post("/", response_model=DNCListResponse, status_code=201)
async def add_to_dnc(
    data: DNCListCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> DNCListResponse:
    """Add a phone number to the DNC list."""
    existing = db.query(DNCList).filter(DNCList.phone_number == data.phone_number).first()
    if existing:
        raise HTTPException(status_code=409, detail="Number already on DNC list")

    entry = DNCList(
        phone_number=data.phone_number,
        source=data.source,
        reason=data.reason,
        added_by=current_user.id,
    )
    db.add(entry)

    # Mark matching dial list entries as DNC
    db.query(DialListEntry).filter(
        DialListEntry.phone_number == data.phone_number
    ).update({"dnc_flag": True, "status": DialEntryStatus.DNC})

    db.commit()
    db.refresh(entry)
    return DNCListResponse.model_validate(entry)


@dnc_router.delete("/{dnc_id}")
async def remove_from_dnc(
    dnc_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Remove a phone number from the DNC list."""
    entry = db.get(DNCList, dnc_id)
    if not entry:
        raise HTTPException(status_code=404, detail="DNC entry not found")

    # Unflag matching dial list entries
    db.query(DialListEntry).filter(
        DialListEntry.phone_number == entry.phone_number,
        DialListEntry.dnc_flag == True,
    ).update({"dnc_flag": False, "status": DialEntryStatus.NEW})

    db.delete(entry)
    db.commit()
    return {"message": "Number removed from DNC list", "id": dnc_id}


@dnc_router.post("/import", response_model=ExcelUploadResponse)
async def import_dnc_file(
    file: UploadFile = File(...),
    phone_column: str = Form("A"),
    reason: str = Form("import"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> ExcelUploadResponse:
    """Bulk import DNC numbers from a CSV or Excel file."""
    filename = (file.filename or "").lower()
    if not filename.endswith((".xlsx", ".xls", ".csv")):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, and .csv files are supported")

    contents = await file.read()
    if len(contents) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="File too large")

    column_mapping = {"phone_column": phone_column}

    try:
        rows = _parse_excel_rows(contents, filename, column_mapping)
    except Exception as e:
        raise HTTPException(status_code=400, detail="Failed to parse file. Check the file format.")

    existing_dnc = set(row[0] for row in db.query(DNCList.phone_number).all())

    success = 0
    duplicates = 0
    errors = 0
    batch: List[dict] = []

    for row_data in rows:
        phone = row_data.get("phone_number")
        if not phone:
            errors += 1
            continue
        if phone in existing_dnc:
            duplicates += 1
            continue

        batch.append({
            "phone_number": phone,
            "source": "import",
            "reason": reason,
            "added_by": current_user.id,
        })
        existing_dnc.add(phone)
        success += 1

        if len(batch) >= 1000:
            db.bulk_insert_mappings(DNCList, batch)
            batch.clear()

    if batch:
        db.bulk_insert_mappings(DNCList, batch)

    # Mark matching dial list entries as DNC
    if success > 0:
        db.query(DialListEntry).filter(
            DialListEntry.phone_number.in_(existing_dnc),
            DialListEntry.dnc_flag == False,
        ).update({"dnc_flag": True, "status": DialEntryStatus.DNC}, synchronize_session=False)

    db.commit()

    return ExcelUploadResponse(
        total=len(rows),
        success=success,
        errors=errors,
        duplicates=duplicates,
    )


@dnc_router.get("/check/{phone_number}")
async def check_dnc(
    phone_number: str,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Check if a phone number is on the DNC list."""
    entry = db.query(DNCList).filter(DNCList.phone_number == phone_number).first()
    return {"phone_number": phone_number, "is_dnc": entry is not None}


# ============================================================
# Campaign Lists (M2M)
# ============================================================

campaign_list_router = APIRouter(prefix="/campaigns", tags=["Campaign Lists"])


@campaign_list_router.get("/{campaign_id}/lists", response_model=List[CampaignListResponse])
async def get_campaign_lists(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[CampaignListResponse]:
    """Get all dial lists linked to a campaign."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    links = (
        db.query(CampaignList)
        .filter(CampaignList.campaign_id == campaign_id)
        .order_by(CampaignList.priority.desc())
        .all()
    )
    return [CampaignListResponse.model_validate(link) for link in links]


@campaign_list_router.post("/{campaign_id}/lists", response_model=CampaignListResponse, status_code=201)
async def link_list_to_campaign(
    campaign_id: int,
    data: CampaignListCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> CampaignListResponse:
    """Assign a dial list to a campaign."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    dial_list = db.get(DialList, data.list_id)
    if not dial_list:
        raise HTTPException(status_code=404, detail="Dial list not found")

    existing = (
        db.query(CampaignList)
        .filter(CampaignList.campaign_id == campaign_id, CampaignList.list_id == data.list_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="List already linked to this campaign")

    link = CampaignList(
        campaign_id=campaign_id,
        list_id=data.list_id,
        priority=data.priority,
        active=data.active,
    )
    db.add(link)
    db.commit()
    db.refresh(link)
    return CampaignListResponse.model_validate(link)


@campaign_list_router.delete("/{campaign_id}/lists/{list_id}")
async def unlink_list_from_campaign(
    campaign_id: int,
    list_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Remove a dial list from a campaign."""
    link = (
        db.query(CampaignList)
        .filter(CampaignList.campaign_id == campaign_id, CampaignList.list_id == list_id)
        .first()
    )
    if not link:
        raise HTTPException(status_code=404, detail="Campaign-list link not found")

    db.delete(link)
    db.commit()
    return {"message": "List unlinked from campaign", "campaign_id": campaign_id, "list_id": list_id}


# ============================================================
# Campaign Dispositions
# ============================================================

disposition_router = APIRouter(prefix="/campaigns", tags=["Dispositions"])


@disposition_router.get("/{campaign_id}/dispositions", response_model=List[CampaignDispositionResponse])
async def list_dispositions(
    campaign_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> List[CampaignDispositionResponse]:
    """List all dispositions for a campaign."""
    items = (
        db.query(CampaignDisposition)
        .filter(CampaignDisposition.campaign_id == campaign_id)
        .order_by(CampaignDisposition.name)
        .all()
    )
    return [CampaignDispositionResponse.model_validate(item) for item in items]


@disposition_router.post("/{campaign_id}/dispositions", response_model=CampaignDispositionResponse, status_code=201)
async def create_disposition(
    campaign_id: int,
    data: CampaignDispositionCreate,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> CampaignDispositionResponse:
    """Create a new campaign disposition."""
    campaign = db.get(Campaign, campaign_id)
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    disposition = CampaignDisposition(
        campaign_id=campaign_id,
        name=data.name,
        category=data.category,
        next_action=data.next_action,
        retry_delay_minutes=data.retry_delay_minutes,
        is_final=data.is_final,
    )
    db.add(disposition)
    db.commit()
    db.refresh(disposition)
    return CampaignDispositionResponse.model_validate(disposition)


@disposition_router.delete("/{campaign_id}/dispositions/{disposition_id}")
async def delete_disposition(
    campaign_id: int,
    disposition_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> dict:
    """Delete a campaign disposition."""
    disposition = (
        db.query(CampaignDisposition)
        .filter(
            CampaignDisposition.id == disposition_id,
            CampaignDisposition.campaign_id == campaign_id,
        )
        .first()
    )
    if not disposition:
        raise HTTPException(status_code=404, detail="Disposition not found")

    db.delete(disposition)
    db.commit()
    return {"message": "Disposition deleted successfully", "id": disposition_id}


# ============================================================
# Dial Hopper
# ============================================================

hopper_router = APIRouter(prefix="/hopper", tags=["Dial Hopper"])


@hopper_router.get("/{campaign_id}", response_model=PaginatedResponse)
async def list_hopper(
    campaign_id: int,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> PaginatedResponse:
    """List hopper entries for a campaign."""
    query = db.query(DialHopper).filter(DialHopper.campaign_id == campaign_id)
    if status:
        query = query.filter(DialHopper.status == status)

    total = query.count()
    items = (
        query.order_by(DialHopper.priority.desc(), DialHopper.inserted_at)
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )

    return PaginatedResponse(
        items=[DialHopperResponse.model_validate(item) for item in items],
        total=total,
        page=page,
        page_size=page_size,
        pages=(total + page_size - 1) // page_size,
    )
